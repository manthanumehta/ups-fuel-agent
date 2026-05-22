import pandas as pd
import os
import smtplib
import time
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image as OpenpyxlImage
import matplotlib.pyplot as plt
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import warnings
from bs4 import BeautifulSoup

# Selenium Imports
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager

warnings.filterwarnings('ignore')

# --- CONFIGURATION ---
UPS_URL = "https://www.ups.com/in/en/support/shipping-support/shipping-costs-rates/fuel-surcharges.page"
EXCEL_FILE = "ups_fuel_history.xlsx"
TEMP_GRAPH = "ups_plot.png"

# --- EMAIL SECRETS ---
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
EMAIL_SENDER = os.environ.get("EMAIL_SENDER")
EMAIL_PASSWORD = os.environ.get("EMAIL_PASSWORD")
EMAIL_RECEIVERS = [EMAIL_SENDER]

def get_date_with_suffix():
    now = datetime.now()
    day = now.day
    suffix = 'th' if 11 <= day <= 13 else {1: 'st', 2: 'nd', 3: 'rd'}.get(day % 10, 'th')
    return now.strftime(f"{day}{suffix} %B %Y")

def extrapolate_and_expand(df):
    col_start, col_till = 'At Least (USD)', 'But Less Than (USD)'
    
    # ---------------------------------------------------------
    # 1. DOWNWARD LOGIC (Calculated ONLY from Top Edge)
    # ---------------------------------------------------------
    price_step_down = round(df.iloc[1][col_start] - df.iloc[0][col_start], 2)
    rate_change_down = round(df.iloc[1]['Surcharge'] - df.iloc[0]['Surcharge'], 2)
    
    down_rows = []
    curr_till = df.iloc[0][col_start] 
    curr_start = round(curr_till - price_step_down, 2)
    curr_sur = round(df.iloc[0]['Surcharge'] - rate_change_down, 2)
    
    while curr_start >= 0.99:
        down_rows.insert(0, {
            col_start: curr_start, col_till: curr_till, 
            'Surcharge': max(0, curr_sur), 'Steps': price_step_down
        })
        curr_till = curr_start
        curr_start = round(curr_till - price_step_down, 2)
        curr_sur = round(curr_sur - rate_change_down, 2)

    # ---------------------------------------------------------
    # 2. UPWARD LOGIC (Calculated ONLY from Bottom Edge)
    # ---------------------------------------------------------
    price_step_up = round(df.iloc[-1][col_start] - df.iloc[-2][col_start], 2)
    rate_change_up = round(df.iloc[-1]['Surcharge'] - df.iloc[-2]['Surcharge'], 2)
    
    up_rows = []
    curr_start = df.iloc[-1][col_till]
    curr_till = round(curr_start + price_step_up, 2)
    curr_sur = round(df.iloc[-1]['Surcharge'] + rate_change_up, 2)
    
    while curr_start < 5.00:
        up_rows.append({
            col_start: curr_start, col_till: curr_till, 
            'Surcharge': curr_sur, 'Steps': price_step_up
        })
        curr_start = curr_till
        curr_till = round(curr_start + price_step_up, 2)
        curr_sur = round(curr_sur + rate_change_up, 2)

    # ---------------------------------------------------------
    # 3. COMBINE AND EXPAND TO 1-CENT STEPS
    # ---------------------------------------------------------
    full_df = pd.concat([pd.DataFrame(down_rows), df, pd.DataFrame(up_rows)], ignore_index=True)
    full_df = full_df[(full_df[col_start] >= 1.00) & (full_df[col_start] < 5.00)]

    exp_rows = []
    for _, row in full_df.iterrows():
        s, t = row[col_start], row[col_till]
        while round(s, 2) < round(t, 2) and round(s, 2) < 5.00:
            exp_rows.append({
                col_start: round(s, 2), col_till: t, 
                'Surcharge': row['Surcharge'], 'Steps': row['Steps']
            })
            s += 0.01
            
    return pd.DataFrame(exp_rows)

def get_live_data():
    print("Launching REAL Headless Chrome Browser...")
    
    chrome_options = Options()
    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("window-size=1920,1080")
    chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")

    try:
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        driver.set_page_load_timeout(60) 
        
        print("Navigating to UPS Website...")
        driver.get(UPS_URL)
        time.sleep(10) 
        
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        driver.quit()
        
        tables = soup.find_all('table')
        target_table = None
        for table in tables:
            if "Gulf Coast" in table.text or "At Least" in table.text:
                target_table = table
                break
        
        if not target_table:
            return None
            
        data = []
        for row in target_table.find_all('tr'):
            cols = [ele.text.strip() for ele in row.find_all(['td', 'th'])]
            if len(cols) >= 3 and any(c.isdigit() for c in cols[0]):
                try:
                    at_least = float(cols[0].replace('$', '').replace('USD', '').strip())
                    less_than = float(cols[1].replace('$', '').replace('USD', '').strip())
                    surcharge = float(cols[2].replace('%', '').replace(',', '.').strip())
                    data.append([at_least, less_than, surcharge])
                except ValueError:
                    continue
                    
        if not data: return None

        df = pd.DataFrame(data, columns=['At Least (USD)', 'But Less Than (USD)', 'Surcharge'])
        df['Steps'] = (df['But Less Than (USD)'] - df['At Least (USD)']).round(2)
        
        return extrapolate_and_expand(df)
            
    except Exception as e:
        if 'driver' in locals(): driver.quit()
        return None

def send_email(subject, body):
    if not EMAIL_SENDER or not EMAIL_PASSWORD:
        return
        
    msg = MIMEMultipart()
    msg['From'] = EMAIL_SENDER
    msg['To'] = ", ".join(EMAIL_RECEIVERS)
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))

    if os.path.exists(EXCEL_FILE):
        with open(EXCEL_FILE, "rb") as f:
            part = MIMEApplication(f.read(), Name=os.path.basename(EXCEL_FILE))
        part['Content-Disposition'] = f'attachment; filename="{os.path.basename(EXCEL_FILE)}"'
        msg.attach(part)

    try:
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(EMAIL_SENDER, EMAIL_PASSWORD)
        server.send_message(msg)
        server.quit()
    except Exception as e:
        pass

def run_agent():
    today_df = get_live_data()
    if today_df is None: 
        return

    today_date_str = get_date_with_suffix()
    changed_prices = set()
    email_body = f"Automated UPS Fuel Surcharge Report for {today_date_str}.\n\n"
    has_changes = False

    if os.path.exists(EXCEL_FILE):
        with pd.ExcelFile(EXCEL_FILE) as xls:
            last_sheet = xls.sheet_names[-1]
            yesterday_df = pd.read_excel(xls, sheet_name=last_sheet)
        
        merged = yesterday_df.merge(today_df, on=['At Least (USD)'], suffixes=('_old', '_new'))
        s_changes = merged[abs(merged['Surcharge_old'] - merged['Surcharge_new']) > 0.001]
        st_changes = merged[abs(merged['Steps_old'] - merged['Steps_new']) > 0.001]
        
        changed_prices.update(s_changes['At Least (USD)'].tolist())
        changed_prices.update(st_changes['At Least (USD)'].tolist())

        if not s_changes.empty or not st_changes.empty:
            has_changes = True
            
            if not s_changes.empty: 
                email_body += f"[ALERT] SURCHARGE RATE CHANGES DETECTED ({len(s_changes)} rows impacted)\n"
                for _, row in s_changes.head(20).iterrows():
                    email_body += f"   > Price Point: ${row['At Least (USD)']} | Rate: {row['Surcharge_old']}% -> {row['Surcharge_new']}%\n"
                email_body += "\n"
                
            if not st_changes.empty: 
                email_body += f"[ALERT] INFLECTION POINT CHANGES DETECTED ({len(st_changes)} rows impacted)\n"
                for _, row in st_changes.head(20).iterrows():
                    email_body += f"   > Price Point: ${row['At Least (USD)']} | Step Size: {row['Steps_old']} -> {row['Steps_new']}\n"
                email_body += "\n"
                
            email_body += "Please see the attached Excel file for the full data. Changed rows are highlighted in YELLOW.\n"
        else:
            email_body += "Status: No changes detected since yesterday.\n"

        # --- UPDATED PLOTTING SECTION ---
        plt.figure(figsize=(10, 6))
        plt.gca().set_facecolor('#FFFEE0')
        
        # 1. PREVIOUS: Thick, semi-transparent shadow
        plt.plot(yesterday_df['At Least (USD)'], yesterday_df['Surcharge'], 
                 color='grey', linewidth=7, alpha=0.4, label='Previous')
        
        # 2. CURRENT: Crisp, solid line on top
        plt.plot(today_df['At Least (USD)'], today_df['Surcharge'], 
                 color='#351C15', linewidth=2, label='Current')
        
        plt.title('UPS Fuel Surcharge Index', fontweight='bold')
        plt.xlabel('Fuel Price (USD)')
        plt.ylabel('Surcharge (%)')
        
        # Add subtle grid
        plt.grid(True, linestyle=':', alpha=0.6) 
        
        plt.legend()
        plt.savefig(TEMP_GRAPH, bbox_inches='tight')
        plt.close()
        # --------------------------------
    else:
        email_body += "Initial baseline established.\n"

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    mode = 'a' if os.path.exists(EXCEL_FILE) else 'w'
    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode=mode if mode == 'a' else 'w', if_sheet_exists='replace' if mode=='a' else None) as writer:
        today_df.to_excel(writer, sheet_name=today_date_str, index=False)
        worksheet = writer.sheets[today_date_str]
        for col_idx in range(1, 5):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = yellow_fill; cell.font = Font(bold=True); cell.alignment = Alignment(horizontal='center')
        for row_idx, start_val in enumerate(today_df['At Least (USD)'], start=2):
            worksheet.cell(row=row_idx, column=3).number_format = '0.00"%"'
            if start_val in changed_prices:
                for col_idx in range(1, 5): worksheet.cell(row=row_idx, column=col_idx).fill = yellow_fill
        if os.path.exists(TEMP_GRAPH):
            worksheet.add_image(OpenpyxlImage(TEMP_GRAPH), 'F2')

    if os.path.exists(TEMP_GRAPH): os.remove(TEMP_GRAPH)
    
    subject = f"UPS Surcharge Alert - {today_date_str}" if has_changes else f"UPS Surcharge Log - {today_date_str}"
    send_email(subject, email_body)

if __name__ == "__main__":
    run_agent()
